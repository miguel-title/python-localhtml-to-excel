#import the datetime in order to get the current time
from datetime import datetime

#import the csv in order to write the data to csv
import csv

#Import the Selenium 
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


from selenium.webdriver.chrome.options import Options

import time
import random

#Import the os in order to get the current directory
import os
ROOT_PATH = os.path.realpath(os.path.dirname(os.path.abspath(__file__)))

class SinglesApp():
    def __init__(self, name=""):
        #self.name: str = name
        self.driver = self.make_driver()

        #Set the Constant
        self.BaseUrl = "https://www.cnbc.com/quotes/"
        
        self.inputpath = ''
        self.outputpath = ''
        self.errorsymbols = []


    def make_driver(self):
        options = Options()
        if os.name == "nt":
            options.add_argument('--headless')
            options.add_argument("--no-sandbox")
            driver = webdriver.Chrome(options=options)
        else:
            options.add_argument('--headless')
            options.add_argument("--no-sandbox")
            driver = webdriver.Chrome(chrome_options=options, executable_path='./chromedriver')


        driver.maximize_window()
        driver.implicitly_wait(10)

        return driver


    def close(self):
    	#Close the selenium driver
        self.driver.quit()


    def time_sleep(self, type):
        if type == 1:
            sleeptime = random.randrange(10,100)/100
        elif type == 2:
            sleeptime = random.randrange(70, 200)/100
        elif type == 3:
            sleeptime = random.randrange(100, 300)/100
        elif type == 4:
            sleeptime = random.randrange(150, 400)/100
        elif type == 5:
            sleeptime = random.randrange(400, 500)/100
        elif type == 401:
            sleeptime = random.randrange(60, 100)
        time.sleep(sleeptime)

    def getInputHtmlsPath(self):
        lstSymbols = []

        import configparser

        parser = configparser.ConfigParser()
        parser.read('setting.ini')

        self.inputpath = parser.get('global', 'input')
        self.outputpath = parser.get('global',  'output')

        import openpyxl

        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook(self.inputpath)

        # Define variable to read the active sheet:
        worksheet = wookbook.active
        
        for i in range(0, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                lstSymbols.append(col[i].value)
                break

        return lstSymbols

    def processing(self):
        lstInputPaths = self.getInputHtmlsPath()

        result = []
        header_data = ['TICKER', 'EPS (TTM)', 'P/E (TTM)', 'FWD P/E (NTM)', 'EBITDA (MRQ)', 'ROE (MRQ)', 'REVENUE (MRQ)', 'GROSS MARGIN (MRQ)', 'NET MARGIN (MRQ)', 'DEBT TO EQUITY (MRQ)']

        result.append(header_data)
        isfirst = True
        index = 1
        for path in lstInputPaths:
            if isfirst == True:
                isfirst = False
                continue
            #print(index)
            index += 1
            print(symbol)
            url = self.BaseUrl + symbol + "?qsearchterm="
            print(url)

            driver = self.driver
            try:
                driver.get(url)
            except Exception as e:
                print('symbol {} error'.format(symbol))
                continue

            try:
                eps = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, "//*[@id='MainContentContainer']/div/div[2]/div[1]/div[5]/div[2]/section/div[3]/ul/li[1]/span[2]"))).get_attribute('textContent')
                pe = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, "//*[@id='MainContentContainer']/div/div[2]/div[1]/div[5]/div[2]/section/div[3]/ul/li[2]/span[2]"))).get_attribute('textContent')
                fwd = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, "//*[@id='MainContentContainer']/div/div[2]/div[1]/div[5]/div[2]/section/div[3]/ul/li[3]/span[2]"))).get_attribute('textContent')
                ebitda = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, "//*[@id='MainContentContainer']/div/div[2]/div[1]/div[5]/div[2]/section/div[3]/ul/li[4]/span[2]"))).get_attribute('textContent')
                roe = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, "//*[@id='MainContentContainer']/div/div[2]/div[1]/div[5]/div[2]/section/div[3]/ul/li[5]/span[2]"))).get_attribute('textContent')
                revenue = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, "//*[@id='MainContentContainer']/div/div[2]/div[1]/div[5]/div[2]/section/div[3]/ul/li[6]/span[2]"))).get_attribute('textContent')
                grossmargin = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, "//*[@id='MainContentContainer']/div/div[2]/div[1]/div[5]/div[2]/section/div[3]/ul/li[7]/span[2]"))).get_attribute('textContent')
                netmargin = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, "//*[@id='MainContentContainer']/div/div[2]/div[1]/div[5]/div[2]/section/div[3]/ul/li[8]/span[2]"))).get_attribute('textContent')
                debttoequity = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH, "//*[@id='MainContentContainer']/div/div[2]/div[1]/div[5]/div[2]/section/div[3]/ul/li[9]/span[2]"))).get_attribute('textContent')

                
                onedata = [symbol, eps, pe, fwd, ebitda, roe, revenue, grossmargin, netmargin, debttoequity]
            except:
                self.errorsymbols.append(symbol)
                continue

            print(onedata)
            print('----------------------')

            result.append(onedata)
            
        self.writeDatatoExcel(result)

        #print(self.errorsymbols)
        return

    def writeDatatoExcel(self, result):
        outputfilename = self.outputpath + '/' + datetime.now().strftime("%Y%m%d%H%M%S") + '.xlsx'
        import xlsxwriter

        workbook = xlsxwriter.Workbook(outputfilename)
        worksheet = workbook.add_worksheet()

        for row_num, row_data in enumerate(result):
            for col_num, col_data in enumerate(row_data):
                worksheet.write(row_num, col_num, col_data)

        workbook.close()


if __name__ == "__main__":
    app = SinglesApp()
    app.processing()
    app.close()